import json
import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk
from tkinter import font as tkFont
from pandas import json_normalize
import numpy as np

class JSONToTabularConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("🔄 JSON to Tabular Converter")
        self.root.geometry("1400x900")
        self.root.configure(bg="#f8f9fa")
        
        # Data storage
        self.json_data = None
        self.flattened_df = None
        
        # Configure styles
        self.setup_styles()
        
        # Create main container
        main_frame = tk.Frame(root, bg="#f8f9fa")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Create interface sections
        self.create_header(main_frame)
        self.create_file_section(main_frame)
        self.create_conversion_section(main_frame)
        self.create_results_section(main_frame)
        self.create_status_bar()

    def setup_styles(self):
        """Configure color scheme and fonts"""
        self.colors = {
            'primary': '#2563eb',
            'secondary': '#64748b',
            'success': '#10b981',
            'warning': '#f59e0b',
            'danger': '#ef4444',
            'light': '#f1f5f9',
            'dark': '#1e293b',
            'white': '#ffffff',
            'accent': '#8b5cf6'
        }
        
        self.fonts = {
            'title': tkFont.Font(family="Arial", size=24, weight="bold"),
            'heading': tkFont.Font(family="Arial", size=16, weight="bold"),
            'subheading': tkFont.Font(family="Arial", size=12, weight="bold"),
            'normal': tkFont.Font(family="Arial", size=10),
            'small': tkFont.Font(family="Arial", size=9)
        }

    def create_header(self, parent):
        """Create the application header"""
        header_frame = tk.Frame(parent, bg=self.colors['primary'], height=80)
        header_frame.pack(fill="x", pady=(0, 20))
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame,
            text="🔄 JSON to Tabular Converter",
            font=self.fonts['title'],
            bg=self.colors['primary'],
            fg=self.colors['white']
        )
        title_label.pack(expand=True)
        
        subtitle_label = tk.Label(
            header_frame,
            text="Transform nested JSON data into structured tabular format",
            font=self.fonts['normal'],
            bg=self.colors['primary'],
            fg=self.colors['light']
        )
        subtitle_label.pack()

    def create_file_section(self, parent):
        """Create file selection section"""
        file_frame = tk.LabelFrame(
            parent,
            text="📁 JSON File Selection",
            font=self.fonts['heading'],
            bg=self.colors['white'],
            fg=self.colors['primary'],
            padx=20,
            pady=15
        )
        file_frame.pack(fill="x", pady=(0, 20))
        
        # File selection controls
        controls_frame = tk.Frame(file_frame, bg=self.colors['white'])
        controls_frame.pack(fill="x", pady=10)
        
        self.file_path_var = tk.StringVar(value="No file selected")
        
        select_btn = tk.Button(
            controls_frame,
            text="📂 Choose JSON File",
            command=self.select_json_file,
            font=self.fonts['subheading'],
            bg=self.colors['primary'],
            fg=self.colors['white'],
            relief="flat",
            padx=20,
            pady=10,
            cursor="hand2"
        )
        select_btn.pack(side="left", padx=(0, 10))
        
        batch_btn = tk.Button(
            controls_frame,
            text="📊 Batch Convert to Excel",
            command=self.batch_convert_to_excel,
            font=self.fonts['normal'],
            bg=self.colors['success'],
            fg=self.colors['white'],
            relief="flat",
            padx=15,
            pady=10,
            cursor="hand2"
        )
        batch_btn.pack(side="left", padx=(0, 15))
        
        file_label = tk.Label(
            controls_frame,
            textvariable=self.file_path_var,
            font=self.fonts['normal'],
            bg=self.colors['white'],
            fg=self.colors['secondary']
        )
        file_label.pack(side="left", fill="x", expand=True)

    def create_conversion_section(self, parent):
        """Create conversion options section"""
        conversion_frame = tk.LabelFrame(
            parent,
            text="⚙️ Conversion Options",
            font=self.fonts['heading'],
            bg=self.colors['white'],
            fg=self.colors['primary'],
            padx=20,
            pady=15
        )
        conversion_frame.pack(fill="x", pady=(0, 20))
        
        # Options frame
        options_frame = tk.Frame(conversion_frame, bg=self.colors['white'])
        options_frame.pack(fill="x", pady=10)
        
        # Conversion options
        self.separator_var = tk.StringVar(value="_")
        self.max_level_var = tk.StringVar(value="")
        self.handle_arrays_var = tk.BooleanVar(value=True)
        self.remove_nulls_var = tk.BooleanVar(value=False)
        
        # Separator option
        sep_frame = tk.Frame(options_frame, bg=self.colors['white'])
        sep_frame.pack(anchor="w", pady=5)
        
        tk.Label(
            sep_frame,
            text="Separator for nested keys:",
            font=self.fonts['normal'],
            bg=self.colors['white']
        ).pack(side="left")
        
        sep_entry = tk.Entry(
            sep_frame,
            textvariable=self.separator_var,
            font=self.fonts['normal'],
            width=5
        )
        sep_entry.pack(side="left", padx=(10, 0))
        
        # Max level option
        level_frame = tk.Frame(options_frame, bg=self.colors['white'])
        level_frame.pack(anchor="w", pady=5)
        
        tk.Label(
            level_frame,
            text="Max nesting level (empty for all):",
            font=self.fonts['normal'],
            bg=self.colors['white']
        ).pack(side="left")
        
        level_entry = tk.Entry(
            level_frame,
            textvariable=self.max_level_var,
            font=self.fonts['normal'],
            width=5
        )
        level_entry.pack(side="left", padx=(10, 0))
        
        # Checkboxes
        tk.Checkbutton(
            options_frame,
            text="Handle arrays as separate rows",
            variable=self.handle_arrays_var,
            font=self.fonts['normal'],
            bg=self.colors['white']
        ).pack(anchor="w", pady=2)
        
        tk.Checkbutton(
            options_frame,
            text="Remove null/empty values",
            variable=self.remove_nulls_var,
            font=self.fonts['normal'],
            bg=self.colors['white']
        ).pack(anchor="w", pady=2)
        
        # Convert button
        convert_btn = tk.Button(
            conversion_frame,
            text="🔄 Convert to Tabular Format",
            command=self.convert_json_to_tabular,
            font=self.fonts['subheading'],
            bg=self.colors['success'],
            fg=self.colors['white'],
            relief="flat",
            padx=30,
            pady=12,
            cursor="hand2"
        )
        convert_btn.pack(pady=(10, 0))

    def create_results_section(self, parent):
        """Create results display section"""
        results_frame = tk.LabelFrame(
            parent,
            text="📊 Conversion Results",
            font=self.fonts['heading'],
            bg=self.colors['white'],
            fg=self.colors['primary'],
            padx=20,
            pady=15
        )
        results_frame.pack(fill="both", expand=True)
        
        # Create notebook for tabbed results
        self.notebook = ttk.Notebook(results_frame)
        self.notebook.pack(fill="both", expand=True, pady=(10, 0))
        
        # Create tabs
        self.create_tabs()

    def create_tabs(self):
        """Create result tabs"""
        # Original JSON tab
        self.json_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(self.json_frame, text="📄 Original JSON")
        
        self.json_text = scrolledtext.ScrolledText(
            self.json_frame,
            font=tkFont.Font(family="Courier", size=9),
            bg=self.colors['light'],
            relief="flat",
            padx=15,
            pady=15
        )
        self.json_text.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Tabular data tab
        self.tabular_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(self.tabular_frame, text="📋 Tabular Data")
        
        self.tabular_text = scrolledtext.ScrolledText(
            self.tabular_frame,
            font=tkFont.Font(family="Courier", size=9),
            bg=self.colors['light'],
            relief="flat",
            padx=15,
            pady=15
        )
        self.tabular_text.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Summary tab
        self.summary_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(self.summary_frame, text="📈 Summary")
        
        self.summary_text = scrolledtext.ScrolledText(
            self.summary_frame,
            font=self.fonts['normal'],
            bg=self.colors['light'],
            relief="flat",
            padx=15,
            pady=15
        )
        self.summary_text.pack(fill="both", expand=True, padx=10, pady=10)

    def create_status_bar(self):
        """Create status bar"""
        self.status_bar = tk.Label(
            self.root,
            text="Ready to convert JSON files",
            relief="sunken",
            anchor="w",
            font=self.fonts['small'],
            bg=self.colors['light'],
            fg=self.colors['secondary']
        )
        self.status_bar.pack(side="bottom", fill="x")

    def update_status(self, message):
        """Update status bar message"""
        self.status_bar.config(text=message)
        self.root.update_idletasks()

    def select_json_file(self):
        """Handle JSON file selection"""
        try:
            file_path = filedialog.askopenfilename(
                title="Select JSON File",
                filetypes=[
                    ("JSON files", "*.json"),
                    ("All files", "*.*")
                ]
            )
            
            if file_path:
                self.file_path_var.set(file_path)
                self.load_json_file(file_path)
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to select file: {str(e)}")

    def load_json_file(self, file_path):
        """Load and display JSON file"""
        try:
            self.update_status("Loading JSON file...")
            
            with open(file_path, 'r', encoding='utf-8') as file:
                self.json_data = json.load(file)
            
            # Store the current file name for export purposes
            self.current_file_name = os.path.basename(file_path)
            
            # Display original JSON
            self.json_text.delete(1.0, tk.END)
            formatted_json = json.dumps(self.json_data, indent=2, ensure_ascii=False)
            self.json_text.insert(tk.END, formatted_json)
            
            self.update_status(f"JSON file loaded successfully: {os.path.basename(file_path)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load JSON file: {str(e)}")
            self.update_status("Error loading JSON file")

    def convert_json_to_tabular(self):
        """Convert JSON to tabular format"""
        if not self.json_data:
            messagebox.showwarning("Warning", "Please select a JSON file first!")
            return
        
        try:
            self.update_status("Converting JSON to tabular format...")
            
            # Get conversion parameters
            separator = self.separator_var.get() or "_"
            max_level = self.max_level_var.get()
            max_level = int(max_level) if max_level.isdigit() else None
            
            # Convert JSON to DataFrame
            if isinstance(self.json_data, list):
                # Handle array of objects
                self.flattened_df = json_normalize(
                    self.json_data,
                    sep=separator,
                    max_level=max_level
                )
            elif isinstance(self.json_data, dict):
                # Handle single object
                self.flattened_df = json_normalize(
                    [self.json_data],
                    sep=separator,
                    max_level=max_level
                )
            else:
                raise ValueError("JSON data must be an object or array of objects")
            
            # Post-processing options
            if self.remove_nulls_var.get():
                self.flattened_df = self.flattened_df.dropna(how='all', axis=1)
                self.flattened_df = self.flattened_df.replace('', np.nan).dropna(how='all', axis=1)
            
            # Display results
            self.display_tabular_data()
            self.display_summary()
            
            # Enable export options
            self.create_export_section()
            
            self.update_status("Conversion completed successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to convert JSON: {str(e)}")
            self.update_status("Error during conversion")

    def display_tabular_data(self):
        """Display the converted tabular data"""
        self.tabular_text.delete(1.0, tk.END)
        
        if self.flattened_df is not None:
            # Display as formatted table
            tabular_str = self.flattened_df.to_string(index=False, max_rows=100)
            self.tabular_text.insert(tk.END, tabular_str)
            
            if len(self.flattened_df) > 100:
                self.tabular_text.insert(tk.END, f"\n\n... and {len(self.flattened_df) - 100} more rows")

    def display_summary(self):
        """Display conversion summary"""
        self.summary_text.delete(1.0, tk.END)
        
        if self.flattened_df is not None:
            summary = [
                "🔄 JSON TO TABULAR CONVERSION SUMMARY",
                "=" * 50,
                "",
                f"📊 Data Shape:",
                f"   Rows: {len(self.flattened_df):,}",
                f"   Columns: {len(self.flattened_df.columns):,}",
                "",
                f"📋 Column Information:",
                f"   Total columns: {len(self.flattened_df.columns)}",
                f"   Data types: {self.flattened_df.dtypes.value_counts().to_dict()}",
                "",
                f"🔍 Data Quality:",
                f"   Missing values: {self.flattened_df.isnull().sum().sum():,}",
                f"   Complete rows: {len(self.flattened_df.dropna()):,}",
                f"   Memory usage: {self.flattened_df.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB",
                "",
                f"📝 Column Names:",
            ]
            
            # Add column names
            for i, col in enumerate(self.flattened_df.columns, 1):
                summary.append(f"   {i:2d}. {col}")
                if i >= 20:  # Limit display
                    summary.append(f"   ... and {len(self.flattened_df.columns) - 20} more columns")
                    break
            
            self.summary_text.insert(tk.END, "\n".join(summary))

    def create_export_section(self):
        """Create export options after successful conversion"""
        if hasattr(self, 'export_frame'):
            return  # Already created
        
        # Find the conversion frame and add export options
        for widget in self.root.winfo_children():
            if isinstance(widget, tk.Frame):
                for child in widget.winfo_children():
                    if isinstance(child, tk.LabelFrame) and "Conversion Options" in child.cget("text"):
                        self.export_frame = tk.Frame(child, bg=self.colors['white'])
                        self.export_frame.pack(fill="x", pady=(15, 0))
                        
                        export_label = tk.Label(
                            self.export_frame,
                            text="📤 Export Options:",
                            font=self.fonts['subheading'],
                            bg=self.colors['white'],
                            fg=self.colors['primary']
                        )
                        export_label.pack(anchor="w", pady=(0, 10))
                        
                        buttons_frame = tk.Frame(self.export_frame, bg=self.colors['white'])
                        buttons_frame.pack(anchor="w")
                        
                        csv_btn = tk.Button(
                            buttons_frame,
                            text="💾 Export as CSV",
                            command=self.export_to_csv,
                            font=self.fonts['normal'],
                            bg=self.colors['accent'],
                            fg=self.colors['white'],
                            relief="flat",
                            padx=15,
                            pady=8,
                            cursor="hand2"
                        )
                        csv_btn.pack(side="left", padx=(0, 10))
                        
                        excel_btn = tk.Button(
                            buttons_frame,
                            text="📊 Export to Excel",
                            command=self.show_excel_export_dialog,
                            font=self.fonts['normal'],
                            bg=self.colors['warning'],
                            fg=self.colors['white'],
                            relief="flat",
                            padx=15,
                            pady=8,
                            cursor="hand2"
                        )
                        excel_btn.pack(side="left", padx=(0, 10))
                    
                        return

    def export_to_csv(self):
        """Export tabular data to CSV"""
        if self.flattened_df is None:
            messagebox.showwarning("Warning", "No data to export!")
            return
        
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                title="Save as CSV"
            )
            
            if file_path:
                self.flattened_df.to_csv(file_path, index=False)
                messagebox.showinfo("Success", f"Data exported successfully to:\n{file_path}")
                self.update_status(f"Exported to CSV: {os.path.basename(file_path)}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export CSV: {str(e)}")

    def export_to_excel(self):
        """Export tabular data to Excel with enhanced formatting"""
        if self.flattened_df is None:
            messagebox.showwarning("Warning", "No data to export!")
            return
        
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save as Excel"
            )
            
            if file_path:
                # Convert any list/dict columns to strings for Excel compatibility
                df_export = self.flattened_df.copy()
                for col in df_export.columns:
                    if df_export[col].dtype == 'object':
                        df_export[col] = df_export[col].apply(lambda x: str(x) if isinstance(x, (list, dict)) else x)
                
                # Create Excel writer object with openpyxl engine for enhanced formatting
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Write main data to 'Data' sheet
                    df_export.to_excel(writer, sheet_name='Data', index=False)
                    
                    # Get the workbook and worksheet objects
                    workbook = writer.book
                    worksheet = writer.sheets['Data']
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Add header formatting
                    try:
                        from openpyxl.styles import Font, PatternFill, Alignment
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_alignment = Alignment(horizontal="center", vertical="center")
                        
                        for cell in worksheet[1]:  # First row (headers)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                    except ImportError:
                        pass  # Skip formatting if styles not available
                    
                    # Create summary sheet with conversion details
                    summary_data = {
                        'Metric': [
                            'Source File',
                            'Total Rows',
                            'Total Columns', 
                            'Missing Values',
                            'Complete Rows',
                            'Memory Usage (MB)',
                            'Conversion Date',
                            'Separator Used'
                        ],
                        'Value': [
                            getattr(self, 'current_file_name', 'Unknown'),
                            len(df_export),
                            len(df_export.columns),
                            df_export.isnull().sum().sum(),
                            len(df_export.dropna()),
                            round(df_export.memory_usage(deep=True).sum() / 1024 / 1024, 2),
                            pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                            self.separator_var.get() or "_"
                        ]
                    }
                    
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Format summary sheet headers
                    try:
                        summary_sheet = writer.sheets['Summary']
                        for cell in summary_sheet[1]:  # Headers
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        
                        # Auto-adjust summary sheet columns
                        for column in summary_sheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = max_length + 2
                            summary_sheet.column_dimensions[column_letter].width = adjusted_width
                    except:
                        pass  # Skip formatting if there are issues
                
                messagebox.showinfo("Success", f"Excel file exported successfully:\n{file_path}\n\nFeatures included:\n• Formatted headers\n• Auto-sized columns\n• Data sheet with converted JSON\n• Summary sheet with conversion details")
                self.update_status(f"Exported to Excel: {os.path.basename(file_path)}")
                
        except ImportError as e:
            messagebox.showerror("Error", "Excel export requires 'openpyxl' package.\nPlease install it using: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export Excel: {str(e)}")

    def export_to_excel_multiple_sheets(self):
        """Export with advanced Excel features - multiple sheets by category"""
        if self.flattened_df is None:
            messagebox.showwarning("Warning", "No data to export!")
            return
        
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save as Advanced Excel (Multiple Sheets)"
            )
            
            if file_path:
                # Convert any list/dict columns to strings for Excel compatibility
                df_export = self.flattened_df.copy()
                for col in df_export.columns:
                    if df_export[col].dtype == 'object':
                        df_export[col] = df_export[col].apply(lambda x: str(x) if isinstance(x, (list, dict)) else x)
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Main data sheet
                    df_export.to_excel(writer, sheet_name='All_Data', index=False)
                    
                    # Try to create separate sheets based on common patterns
                    # This is useful for nested JSON with different entity types
                    
                    # Look for columns that might represent different entity types
                    potential_categories = []
                    for col in df_export.columns:
                        if any(keyword in col.lower() for keyword in ['type', 'category', 'kind', 'class', 'department']):
                            if df_export[col].nunique() <= 10:  # Reasonable number of categories
                                potential_categories.append(col)
                    
                    # Create sheets by category if found
                    sheets_created = ['All_Data']
                    if potential_categories:
                        category_col = potential_categories[0]  # Use first suitable column
                        for category in df_export[category_col].unique():
                            if pd.notna(category) and str(category).strip():
                                category_data = df_export[df_export[category_col] == category]
                                # Clean sheet name for Excel compatibility
                                sheet_name = str(category).replace('/', '_').replace('\\', '_').replace('[', '').replace(']', '')[:31]
                                if sheet_name not in sheets_created:
                                    category_data.to_excel(writer, sheet_name=sheet_name, index=False)
                                    sheets_created.append(sheet_name)
                    
                    # Column analysis sheet
                    col_analysis = []
                    for col in df_export.columns:
                        col_info = {
                            'Column_Name': col,
                            'Data_Type': str(df_export[col].dtype),
                            'Non_Null_Count': df_export[col].count(),
                            'Null_Count': df_export[col].isnull().sum(),
                            'Null_Percentage': round((df_export[col].isnull().sum() / len(df_export)) * 100, 2),
                            'Unique_Values': df_export[col].nunique(),
                            'Sample_Value': str(df_export[col].dropna().iloc[0]) if not df_export[col].dropna().empty else 'N/A'
                        }
                        col_analysis.append(col_info)
                    
                    analysis_df = pd.DataFrame(col_analysis)
                    analysis_df.to_excel(writer, sheet_name='Column_Analysis', index=False)
                    
                    # Summary sheet
                    summary_data = {
                        'Metric': [
                            'Source File',
                            'Total Rows',
                            'Total Columns', 
                            'Missing Values',
                            'Complete Rows',
                            'Memory Usage (MB)',
                            'Sheets Created',
                            'Category Column Used',
                            'Conversion Date',
                            'Separator Used'
                        ],
                        'Value': [
                            getattr(self, 'current_file_name', 'Unknown'),
                            len(df_export),
                            len(df_export.columns),
                            df_export.isnull().sum().sum(),
                            len(df_export.dropna()),
                            round(df_export.memory_usage(deep=True).sum() / 1024 / 1024, 2),
                            len(sheets_created),
                            potential_categories[0] if potential_categories else 'None',
                            pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                            self.separator_var.get() or "_"
                        ]
                    }
                    
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Apply formatting to all sheets
                    try:
                        from openpyxl.styles import Font, PatternFill, Alignment
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_alignment = Alignment(horizontal="center", vertical="center")
                        
                        for sheet_name in writer.sheets:
                            worksheet = writer.sheets[sheet_name]
                            # Format headers
                            for cell in worksheet[1]:
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                            
                            # Auto-adjust column widths
                            for column in worksheet.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                worksheet.column_dimensions[column_letter].width = adjusted_width
                    except ImportError:
                        pass  # Skip formatting if not available
                
                category_info = f"\n• {len(sheets_created)} sheets created" if len(sheets_created) > 3 else "\n• Standard sheets only"
                messagebox.showinfo("Success", f"Advanced Excel file created:\n{file_path}\n\nIncludes:\n• All data sheet\n• Column analysis sheet\n• Summary sheet{category_info}")
                self.update_status(f"Advanced Excel export: {os.path.basename(file_path)}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export advanced Excel: {str(e)}")

    def show_excel_export_dialog(self):
        """Show dialog to choose Excel export type"""
        if self.flattened_df is None:
            messagebox.showwarning("Warning", "No data to export!")
            return
        
        # Create custom dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Excel Export Options")
        dialog.geometry("400x300")
        dialog.configure(bg=self.colors['white'])
        dialog.resizable(False, False)
        
        # Center the dialog
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Main frame
        main_frame = tk.Frame(dialog, bg=self.colors['white'], padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(
            main_frame,
            text="📊 Choose Excel Export Type",
            font=self.fonts['subheading'],
            bg=self.colors['white'],
            fg=self.colors['primary']
        )
        title_label.pack(pady=(0, 20))
        
        # Option 1: Basic Excel
        basic_frame = tk.Frame(main_frame, bg=self.colors['light'], relief="ridge", bd=1)
        basic_frame.pack(fill="x", pady=(0, 10))
        
        basic_title = tk.Label(
            basic_frame,
            text="📄 Basic Excel Export",
            font=self.fonts['normal'],
            bg=self.colors['light'],
            fg=self.colors['dark']
        )
        basic_title.pack(pady=(10, 5))
        
        basic_desc = tk.Label(
            basic_frame,
            text="• Single data sheet with formatted headers\n• Summary sheet with conversion details\n• Auto-sized columns\n• Professional formatting",
            font=("Arial", 9),
            bg=self.colors['light'],
            fg=self.colors['secondary'],
            justify="left"
        )
        basic_desc.pack(pady=(0, 10))
        
        basic_btn = tk.Button(
            basic_frame,
            text="Export Basic Excel",
            command=lambda: self._execute_export(dialog, 'basic'),
            font=self.fonts['normal'],
            bg=self.colors['primary'],
            fg=self.colors['white'],
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        basic_btn.pack(pady=(0, 10))
        
        # Option 2: Advanced Excel
        advanced_frame = tk.Frame(main_frame, bg=self.colors['light'], relief="ridge", bd=1)
        advanced_frame.pack(fill="x", pady=(0, 10))
        
        advanced_title = tk.Label(
            advanced_frame,
            text="📈 Advanced Excel Export",
            font=self.fonts['normal'],
            bg=self.colors['light'],
            fg=self.colors['dark']
        )
        advanced_title.pack(pady=(10, 5))
        
        advanced_desc = tk.Label(
            advanced_frame,
            text="• Multiple sheets (data, analysis, summary)\n• Automatic category-based sheet splitting\n• Detailed column analysis\n• Enhanced data insights",
            font=("Arial", 9),
            bg=self.colors['light'],
            fg=self.colors['secondary'],
            justify="left"
        )
        advanced_desc.pack(pady=(0, 10))
        
        advanced_btn = tk.Button(
            advanced_frame,
            text="Export Advanced Excel",
            command=lambda: self._execute_export(dialog, 'advanced'),
            font=self.fonts['normal'],
            bg=self.colors['success'],
            fg=self.colors['white'],
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        advanced_btn.pack(pady=(0, 10))
        
        # Cancel button
        cancel_btn = tk.Button(
            main_frame,
            text="Cancel",
            command=dialog.destroy,
            font=self.fonts['normal'],
            bg=self.colors['secondary'],
            fg=self.colors['white'],
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        cancel_btn.pack(pady=10)
    
    def _execute_export(self, dialog, export_type):
        """Execute the chosen export type"""
        dialog.destroy()
        if export_type == 'basic':
            self.export_to_excel()
        elif export_type == 'advanced':
            self.export_to_excel_multiple_sheets()

    def batch_convert_to_excel(self):
        """Convert multiple JSON files to Excel in batch"""
        try:
            # Select multiple JSON files
            file_paths = filedialog.askopenfilenames(
                title="Select JSON files to convert",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )
            
            if not file_paths:
                return
            
            # Select output directory
            output_dir = filedialog.askdirectory(title="Select output directory for Excel files")
            if not output_dir:
                return
            
            # Create progress dialog
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Batch Processing")
            progress_window.geometry("400x200")
            progress_window.configure(bg=self.colors['white'])
            progress_window.resizable(False, False)
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            # Progress widgets
            progress_frame = tk.Frame(progress_window, bg=self.colors['white'], padx=20, pady=20)
            progress_frame.pack(fill="both", expand=True)
            
            progress_label = tk.Label(
                progress_frame,
                text="Processing JSON files...",
                font=self.fonts['normal'],
                bg=self.colors['white']
            )
            progress_label.pack(pady=10)
            
            progress_var = tk.StringVar()
            progress_detail = tk.Label(
                progress_frame,
                textvariable=progress_var,
                font=("Arial", 9),
                bg=self.colors['white'],
                fg=self.colors['secondary']
            )
            progress_detail.pack(pady=5)
            
            progress_bar = ttk.Progressbar(
                progress_frame,
                length=300,
                mode='determinate',
                maximum=len(file_paths)
            )
            progress_bar.pack(pady=10)
            
            # Process files
            successful = 0
            failed = 0
            
            for i, file_path in enumerate(file_paths):
                filename = os.path.basename(file_path)
                progress_var.set(f"Processing: {filename}")
                progress_window.update()
                
                try:
                    # Load JSON
                    with open(file_path, 'r', encoding='utf-8') as file:
                        json_data = json.load(file)
                    
                    # Convert to DataFrame
                    if isinstance(json_data, list):
                        df = json_normalize(json_data, sep="_")
                    elif isinstance(json_data, dict):
                        df = json_normalize([json_data], sep="_")
                    else:
                        raise ValueError("Invalid JSON structure")
                    
                    # Convert any list/dict columns to strings
                    for col in df.columns:
                        if df[col].dtype == 'object':
                            df[col] = df[col].apply(lambda x: str(x) if isinstance(x, (list, dict)) else x)
                    
                    # Generate output filename
                    base_name = os.path.splitext(filename)[0]
                    output_file = os.path.join(output_dir, f"{base_name}_converted.xlsx")
                    
                    # Export to Excel
                    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='Data', index=False)
                        
                        # Add summary
                        summary_data = {
                            'Metric': ['Source File', 'Rows', 'Columns', 'Conversion Date'],
                            'Value': [filename, len(df), len(df.columns), pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')]
                        }
                        summary_df = pd.DataFrame(summary_data)
                        summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    successful += 1
                    
                except Exception as e:
                    print(f"Failed to process {filename}: {str(e)}")
                    failed += 1
                
                progress_bar['value'] = i + 1
                progress_window.update()
            
            # Show results
            progress_window.destroy()
            result_msg = f"Batch conversion completed!\n\n✅ Successful: {successful} files\n❌ Failed: {failed} files\n\n📁 Output directory: {output_dir}"
            messagebox.showinfo("Batch Conversion Complete", result_msg)
            
        except Exception as e:
            messagebox.showerror("Error", f"Batch conversion failed: {str(e)}")


if __name__ == "__main__":
    print("🔄 Starting JSON to Tabular Converter...")
    root = tk.Tk()
    app = JSONToTabularConverter(root)
    print("✅ Application initialized. Opening GUI window...")
    root.mainloop()
    print("👋 Application closed.")
